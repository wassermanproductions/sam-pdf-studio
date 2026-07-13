#!/usr/bin/env python3
from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
ENGINE = ROOT / "Engine" / "pdf_engine.py"
TMP = ROOT / "tmp" / "pdfs" / "qa"


def run_engine(*args: str) -> dict:
    env = os.environ.copy()
    env["PATH"] = ":".join([
        str(Path(sys.executable).resolve().parent),
        "/opt/homebrew/bin",
        "/usr/local/bin",
        env.get("PATH", ""),
    ])
    font_config = Path("/opt/homebrew/etc/fonts/fonts.conf")
    if font_config.exists():
        env.setdefault("FONTCONFIG_FILE", str(font_config))
    result = subprocess.run([sys.executable, str(ENGINE), *args], text=True, capture_output=True, env=env)
    try:
        payload = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise AssertionError(f"Engine did not return JSON for {args}: {exc}\nSTDOUT:\n{result.stdout}\nSTDERR:\n{result.stderr}") from exc
    if result.returncode != 0 or not payload.get("ok"):
        raise AssertionError(f"Engine command failed {args}\n{json.dumps(payload, indent=2)}\nSTDERR:\n{result.stderr}")
    return payload


def make_pdf(path: Path, title: str, lines: list[str]) -> None:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    c = canvas.Canvas(str(path), pagesize=letter)
    width, height = letter
    c.setFont("Helvetica-Bold", 18)
    c.drawString(72, height - 72, title)
    c.setFont("Helvetica", 12)
    y = height - 112
    for line in lines:
        c.drawString(72, y, line)
        y -= 22
    c.save()


def make_scanned_pdf(path: Path) -> None:
    from PIL import Image, ImageDraw, ImageFont

    image = Image.new("RGB", (1800, 1200), "white")
    draw = ImageDraw.Draw(image)
    font_paths = [
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Helvetica.ttf",
        "/Library/Fonts/Arial.ttf",
    ]
    font = None
    for font_path in font_paths:
        if Path(font_path).exists():
            font = ImageFont.truetype(font_path, 88)
            break
    if font is None:
        font = ImageFont.load_default()
    draw.text((160, 220), "OCR PRIVATE TEST 2468", fill="black", font=font)
    draw.text((160, 360), "SEARCHABLE PDF CHECK", fill="black", font=font)
    image.save(path, "PDF", resolution=300)


def make_image(path: Path, text: str) -> None:
    from PIL import Image, ImageDraw, ImageFont

    image = Image.new("RGB", (900, 420), "#ffffff")
    draw = ImageDraw.Draw(image)
    font_paths = [
        "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Helvetica.ttf",
        "/Library/Fonts/Arial.ttf",
    ]
    font = None
    for font_path in font_paths:
        if Path(font_path).exists():
            font = ImageFont.truetype(font_path, 52)
            break
    if font is None:
        font = ImageFont.load_default()
    draw.rectangle((40, 40, 860, 380), outline="#111111", width=6)
    draw.text((90, 165), text, fill="#111111", font=font)
    image.save(path)


def extract_text(path: Path) -> str:
    import fitz

    doc = fitz.open(path)
    text = "\n".join(page.get_text("text") for page in doc)
    doc.close()
    return text


def page_text(path: Path, page_index: int) -> str:
    import fitz

    doc = fitz.open(path)
    text = doc[page_index].get_text("text")
    doc.close()
    return text


def page_count(path: Path) -> int:
    import fitz

    doc = fitz.open(path)
    count = doc.page_count
    doc.close()
    return count


def annotation_count(path: Path) -> int:
    import fitz

    doc = fitz.open(path)
    count = 0
    for page in doc:
        annot = page.first_annot
        while annot:
            count += 1
            annot = annot.next
    doc.close()
    return count


def link_count(path: Path) -> int:
    import fitz

    doc = fitz.open(path)
    count = sum(len(page.get_links()) for page in doc)
    doc.close()
    return count


def image_count(path: Path) -> int:
    import fitz

    doc = fitz.open(path)
    count = sum(len(page.get_images(full=True)) for page in doc)
    doc.close()
    return count


def first_search_rect(path: Path, text: str) -> tuple[int, str]:
    import fitz

    doc = fitz.open(path)
    for index, page in enumerate(doc, start=1):
        rects = page.search_for(text)
        if rects:
            rect = rects[0]
            doc.close()
            return index, f"{rect.x0},{rect.y0},{rect.x1},{rect.y1}"
    doc.close()
    raise AssertionError(f"No rect found for {text}")


def vertically_mirrored_rect(path: Path, page_number: int, rect_text: str) -> str:
    import fitz

    doc = fitz.open(path)
    page = doc[page_number - 1]
    x0, y0, x1, y1 = [float(part) for part in rect_text.split(",")]
    height = page.rect.height
    doc.close()
    return f"{x0},{height - y1},{x1},{height - y0}"


def assert_file(path: Path, min_size: int = 1) -> None:
    assert path.exists(), f"Missing file: {path}"
    assert path.stat().st_size >= min_size, f"File too small: {path}"


def main() -> int:
    if TMP.exists():
        shutil.rmtree(TMP)
    TMP.mkdir(parents=True)

    first = TMP / "first.pdf"
    second = TMP / "second.pdf"
    make_pdf(first, "QA First PDF", ["Invoice Total: $100", "ReplaceMe", "This is the first source PDF."])
    make_pdf(second, "QA Second PDF", ["Second file line", "This validates merge order."])

    health = run_engine("health")
    assert health["packages"]["pymupdf"]["ok"]
    assert health["packages"]["pdf2docx"]["ok"]
    assert health["packages"]["ocrmypdf"]["ok"]
    assert health["packages"]["openpyxl"]["ok"]
    assert health["packages"]["python-pptx"]["ok"]
    assert health["binaries"]["qpdf"]["ok"]
    assert health["binaries"]["tesseract"]["ok"]
    assert health["binaries"]["gs"]["ok"]

    merged = TMP / "merged.pdf"
    payload = run_engine("merge", "--input", str(first), "--input", str(second), "--output", str(merged))
    assert payload["pages"] == 2
    assert_file(merged, 1000)
    assert page_count(merged) == 2

    ordered_merged = TMP / "ordered-merged.pdf"
    payload = run_engine(
        "merge-pages",
        "--page-item", f"{second}::1",
        "--page-item", f"{first}::1",
        "--output", str(ordered_merged),
    )
    assert payload["pages"] == 2
    assert page_count(ordered_merged) == 2
    assert "QA Second PDF" in page_text(ordered_merged, 0)
    assert "QA First PDF" in page_text(ordered_merged, 1)

    split_dir = TMP / "split"
    payload = run_engine("split", "--input", str(merged), "--output-dir", str(split_dir))
    assert payload["pages"] == 2
    for output in payload["outputs"]:
        assert_file(Path(output), 900)

    extracted = TMP / "extracted.pdf"
    run_engine("extract-pages", "--input", str(merged), "--output", str(extracted), "--pages", "2")
    assert page_count(extracted) == 1
    assert "QA Second PDF" in extract_text(extracted)

    deleted = TMP / "deleted.pdf"
    run_engine("delete-pages", "--input", str(merged), "--output", str(deleted), "--pages", "2")
    assert page_count(deleted) == 1
    assert "QA Second PDF" not in extract_text(deleted)

    rotated = TMP / "rotated.pdf"
    run_engine("rotate-pages", "--input", str(merged), "--output", str(rotated), "--pages", "1", "--degrees", "90")
    assert_file(rotated, 1000)

    cropped = TMP / "cropped.pdf"
    run_engine("crop-pages", "--input", str(merged), "--output", str(cropped), "--pages", "1", "--left", "12", "--top", "12", "--right", "12", "--bottom", "12")
    assert_file(cropped, 1000)

    replaced = TMP / "replaced.pdf"
    payload = run_engine(
        "replace-text",
        "--input", str(first),
        "--output", str(replaced),
        "--find", "ReplaceMe",
        "--replace", "ChangedPrivateText",
        "--auto-size",
    )
    assert payload["matches"] == 1
    text = extract_text(replaced)
    assert "ChangedPrivateText" in text
    assert "ReplaceMe" not in text

    rect_page, rect_text = first_search_rect(first, "ReplaceMe")
    rect_replaced = TMP / "rect-replaced.pdf"
    payload = run_engine(
        "replace-text",
        "--input", str(first),
        "--output", str(rect_replaced),
        "--find", "ReplaceMe",
        "--replace", "ClickedEditText",
        "--page", str(rect_page),
        "--rect", rect_text,
        "--auto-size",
        "--match-style",
    )
    assert payload["matches"] == 1
    rect_replaced_text = extract_text(rect_replaced)
    assert "ClickedEditText" in rect_replaced_text
    assert "ReplaceMe" not in rect_replaced_text

    mirrored_rect_replaced = TMP / "mirrored-rect-replaced.pdf"
    payload = run_engine(
        "replace-text",
        "--input", str(first),
        "--output", str(mirrored_rect_replaced),
        "--find", "ReplaceMe",
        "--replace", "FlippedClickEditText",
        "--page", str(rect_page),
        "--rect", vertically_mirrored_rect(first, rect_page, rect_text),
        "--auto-size",
        "--match-style",
    )
    assert payload["matches"] == 1
    mirrored_replaced_text = extract_text(mirrored_rect_replaced)
    assert "FlippedClickEditText" in mirrored_replaced_text
    assert "ReplaceMe" not in mirrored_replaced_text

    added = TMP / "added.pdf"
    run_engine(
        "add-text",
        "--input", str(first),
        "--output", str(added),
        "--page", "1",
        "--x", "72",
        "--y", "260",
        "--text", "AddedPrivateOverlay",
        "--font-size", "14",
    )
    assert "AddedPrivateOverlay" in extract_text(added)

    redacted = TMP / "redacted.pdf"
    payload = run_engine(
        "redact-text",
        "--input", str(first),
        "--output", str(redacted),
        "--find", "Invoice Total",
    )
    assert payload["matches"] == 1
    redacted_text = extract_text(redacted)
    assert "Invoice Total" not in redacted_text
    assert "REDACTED" not in redacted_text

    annotated = TMP / "annotated.pdf"
    payload = run_engine(
        "annotate-text",
        "--input", str(first),
        "--output", str(annotated),
        "--kind", "highlight",
        "--find", "Invoice Total",
        "--note", "Check value",
    )
    assert payload["matches"] == 1
    assert annotation_count(annotated) >= 1

    note_pdf = TMP / "note.pdf"
    run_engine("add-note", "--input", str(first), "--output", str(note_pdf), "--page", "1", "--x", "120", "--y", "160", "--text", "QA note")
    assert annotation_count(note_pdf) >= 1

    signed = TMP / "signed.pdf"
    run_engine("add-signature", "--input", str(first), "--output", str(signed), "--page", "1", "--x", "72", "--y", "310", "--width", "220", "--height", "52", "--text", "Sam QA Signature")
    assert "Sam QA Signature" in extract_text(signed)

    insert_image = TMP / "insert.png"
    make_image(insert_image, "INSERT IMAGE QA")
    image_pdf = TMP / "image-inserted.pdf"
    run_engine("add-image", "--input", str(first), "--output", str(image_pdf), "--image", str(insert_image), "--page", "1", "--x", "300", "--y", "240", "--width", "180", "--height", "84")
    assert_file(image_pdf, 1800)

    paste_page, paste_rect = first_search_rect(first, "ReplaceMe")
    pasted_region = TMP / "region-pasted.pdf"
    payload = run_engine(
        "paste-region",
        "--input", str(first),
        "--source", str(first),
        "--output", str(pasted_region),
        "--source-page", str(paste_page),
        "--source-rect", paste_rect,
        "--destination-page", str(paste_page),
        "--destination-x", "260",
        "--destination-y", "260",
    )
    assert payload["destination_page"] == paste_page
    assert_file(pasted_region, 1800)
    assert "ReplaceMe" in extract_text(pasted_region)
    # The paste must be live vector content, not a rasterized snapshot:
    # the text appears twice (original + pasted copy) and stays searchable.
    import fitz as _fitz

    with _fitz.open(str(pasted_region)) as pasted_doc:
        paste_hits = pasted_doc[paste_page - 1].search_for("ReplaceMe")
    assert len(paste_hits) >= 2, f"pasted region is not live text (search hits: {len(paste_hits)})"

    moved_region = TMP / "region-moved.pdf"
    payload = run_engine(
        "paste-region",
        "--input", str(first),
        "--source", str(first),
        "--output", str(moved_region),
        "--source-page", str(paste_page),
        "--source-rect", paste_rect,
        "--destination-page", str(paste_page),
        "--destination-x", "260",
        "--destination-y", "300",
        "--erase-source",
    )
    assert payload["erased"]
    assert_file(moved_region, 1800)
    # A move keeps the text live: gone from the source spot, searchable
    # exactly once at the destination.
    with _fitz.open(str(moved_region)) as moved_doc:
        move_hits = moved_doc[paste_page - 1].search_for("ReplaceMe")
    assert len(move_hits) == 1, f"moved text should exist exactly once (search hits: {len(move_hits)})"
    original_rect = [float(v) for v in paste_rect.split(",")]
    moved_hit = move_hits[0]
    displacement = abs(moved_hit.x0 - original_rect[0]) + abs(moved_hit.y0 - original_rect[1])
    assert displacement > 20, f"moved text did not actually move (displacement {displacement:.1f})"

    linked = TMP / "linked.pdf"
    run_engine("add-link", "--input", str(first), "--output", str(linked), "--page", "1", "--x", "72", "--y", "92", "--width", "240", "--height", "28", "--url", "https://example.com")
    assert link_count(linked) >= 1

    image_dir = TMP / "images"
    payload = run_engine("export-images", "--input", str(merged), "--output-dir", str(image_dir), "--format", "png", "--dpi", "160")
    assert payload["pages"] == 2
    for output in payload["outputs"]:
        assert_file(Path(output), 1000)

    text_out = TMP / "merged.txt"
    run_engine("export-text", "--input", str(merged), "--output", str(text_out), "--format", "txt")
    assert "QA First PDF" in text_out.read_text(encoding="utf-8")

    md_out = TMP / "merged.md"
    run_engine("export-text", "--input", str(merged), "--output", str(md_out), "--format", "md")
    assert "QA First PDF" in md_out.read_text(encoding="utf-8")

    html_out = TMP / "merged.html"
    run_engine("export-text", "--input", str(merged), "--output", str(html_out), "--format", "html")
    assert "<html" in html_out.read_text(encoding="utf-8").lower()

    docx_out = TMP / "merged.docx"
    run_engine("export-docx", "--input", str(merged), "--output", str(docx_out))
    assert_file(docx_out, 1000)
    with zipfile.ZipFile(docx_out) as archive:
        xml = archive.read("word/document.xml").decode("utf-8", errors="ignore")
        assert "QA First PDF" in xml

    xlsx_out = TMP / "merged.xlsx"
    run_engine("export-xlsx", "--input", str(merged), "--output", str(xlsx_out))
    assert_file(xlsx_out, 1000)
    from openpyxl import load_workbook

    workbook = load_workbook(xlsx_out, read_only=True)
    sheet = workbook.active
    values = " ".join(str(cell.value or "") for row in sheet.iter_rows() for cell in row)
    workbook.close()
    assert "QA First PDF" in values

    pptx_out = TMP / "merged.pptx"
    run_engine("export-pptx", "--input", str(merged), "--output", str(pptx_out), "--dpi", "120")
    assert_file(pptx_out, 1000)
    with zipfile.ZipFile(pptx_out) as archive:
        assert "ppt/presentation.xml" in archive.namelist()

    image_pdf_out = TMP / "images-to-pdf.pdf"
    second_image = TMP / "second-image.png"
    make_image(second_image, "SECOND IMAGE QA")
    run_engine("images-to-pdf", "--input", str(insert_image), "--input", str(second_image), "--output", str(image_pdf_out))
    assert page_count(image_pdf_out) == 2

    scanned = TMP / "scanned.pdf"
    scanned_ocr = TMP / "scanned-ocr.pdf"
    make_scanned_pdf(scanned)
    run_engine("ocr", "--input", str(scanned), "--output", str(scanned_ocr), "--language", "eng", "--force")
    ocr_text = extract_text(scanned_ocr).upper()
    assert "OCR" in ocr_text and "2468" in ocr_text

    scanned_page_ocr = TMP / "scanned-page-ocr.pdf"
    run_engine("ocr", "--input", str(scanned), "--output", str(scanned_page_ocr), "--language", "eng", "--page", "1")
    page_ocr_text = extract_text(scanned_page_ocr).upper()
    assert "OCR" in page_ocr_text and "2468" in page_ocr_text

    enhanced = TMP / "scanned-enhanced.pdf"
    run_engine("enhance-scan", "--input", str(scanned), "--output", str(enhanced), "--grayscale", "--denoise", "--contrast", "1.4", "--sharpness", "1.2")
    assert_file(enhanced, 1000)

    rendered = TMP / "render-check"
    pdftoppm = shutil.which("pdftoppm") or "/opt/homebrew/bin/pdftoppm"
    render_env = os.environ.copy()
    font_config = Path("/opt/homebrew/etc/fonts/fonts.conf")
    if font_config.exists():
        render_env.setdefault("FONTCONFIG_FILE", str(font_config))
    subprocess.run([pdftoppm, "-png", "-f", "1", "-singlefile", str(merged), str(rendered)], check=True, env=render_env)
    assert_file(TMP / "render-check.png", 1000)

    print(json.dumps({
        "ok": True,
        "qa_dir": str(TMP),
        "checks": [
            "health",
            "merge",
            "merge_ordered_pages",
            "split",
            "extract_pages",
            "delete_pages",
            "rotate_pages",
            "crop_pages",
            "replace_text",
            "add_text",
            "redact_text",
            "annotate_text",
            "add_note",
            "add_signature",
            "add_image",
            "paste_region",
            "move_region",
            "add_link",
            "export_images",
            "export_text",
            "export_markdown",
            "export_html",
            "export_docx",
            "export_xlsx",
            "export_pptx",
            "images_to_pdf",
            "ocr",
            "ocr_current_page",
            "enhance_scan",
            "render_check",
        ],
    }, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
